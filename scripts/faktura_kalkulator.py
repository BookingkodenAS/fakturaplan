#!/usr/bin/env python3
"""
faktura_kalkulator.py  –  We4you AS sentralbord
================================================
Les Zisson Interact månedsstatistikk (Excel) og beregn fakturabeløp
per kunde basert på prismodellene i Fakturaplan-2026.

Bruk:
    python3 faktura_kalkulator.py <sti-til-zisson.xlsx> [--emails-compass 1776] [--emails-fit4 0]

Uten argumenter brukes standardstien til siste Zisson-fil i Fakturaplan-mappen.

Output:
    Fakturagrunnlag_<måned>.xlsx  –  i C:\\Users\\hei\\Claude\\Projects\\Fakturaplan\\
"""

import sys, os, re
from datetime import date

# ── Pristabeller ─────────────────────────────────────────────────────────────

# (fra til pris)  –  STANDARDTABELL 2026
STDTAB = [
    (0,   99,   2960),
    (100, 149,  3580),
    (150, 199,  4090),
    (200, 249,  4670),
    (250, 299,  5760),
    (300, 349,  6830),
    (350, 399,  7920),
    (400, 449,  8980),
    (450, 499, 10080),
    (500, 549, 11150),
    (550, 599, 12200),
    (600, 649, 13260),
    (650, 699, 14360),
    (700, 749, 15430),
    (750, 799, 16530),
    (800, 849, 17590),
    (850, 899, 18680),
    (900, 949, 19750),
    (950,  999,  20820),
    (1000, 1099, 23000),
    (1100, 1199, 24100),
    (1200, 1299, 26260),
    (1300, 1399, 28400),
    (1400, 1499, 30240),
    (1500, 1599, 32710),
    (1600, 1699, 34890),
    (1700, 1799, 37060),
    (1800, 1899, 39200),
    (1900, 1999, 41380),
    (2000, 2099, 43500),
    (2100, 2199, 45650),
    (2200, 2299, 47850),
    (2300, 2399, 49980),
    (2400, 2499, 52160),
]

# CEG-tabell 2026
CEGTAB = [
    (0,   69,  1500),
    (70,  89,  2280),
    (90,  109, 2870),
    (110, 129, 3260),
    (130, 149, 3450),
    (150, 169, 3810),
    (170, 189, 4020),
    (190, 209, 4260),
    (210, 229, 4490),
    (230, 249, 4860),
    (250, 269, 5280),
    (270, 289, 5690),
    (290, 309, 6100),
    (310, 329, 6520),
    (330, 349, 6940),
    (350, 369, 7370),
    (370, 389, 7790),
    (390, 409, 8200),
    (410, 429, 8610),
    (430, 449, 9020),
    (450, 469, 9440),
    (470, 489, 9860),
    (490, 509, 10280),
]

# Caverion 2026 (indeksjustert +3.2%)
CAVERION_TAB = [
    (0,   99,  2885),
    (100, 149, 3303),
    (150, 199, 3756),
    (200, 249, 4314),
    (250, 299, 5315),
    (300, 349, 6275),
    (350, 399, 7306),
]

# Coromatic 2026 – volumtabell (fra Volumtabell Coromatic.xlsx)
# Over 599 anrop: 11 228 + (anrop – 599) × 18 kr/anrop
COROMATIC_TAB = [
    (0,   99,  2885),
    (100, 149, 3303),
    (150, 199, 3756),
    (200, 249, 4314),
    (250, 299, 5315),
    (300, 349, 6275),
    (350, 399, 7306),
    (400, 449, 8308),
    (450, 499, 9278),
    (500, 549, 10392),
    (550, 599, 11228),
]
COROMATIC_OVER_RATE = 18   # kr per anrop over 599

# EFAS/Eurofusion volumtabell 2026
EFAS_TAB = [
    (0,  30, 1330),
    (31, 40, 1540),
    (41, 50, 1780),
    (51, 60, 1980),
    (61, 70, 2200),
    (71, 80, 2495),
    (81, 100, 2850),
]

# ENRX 2026
ENRX_TAB = [
    (0,   99,  2580),
    (100, 149, 3100),
    (150, 199, 3600),
    (200, 249, 4150),
    (250, 299, 4640),
]
ENRX_OVER = 32  # per anrop over 299

# Promon 2025/26
PROMON_TAB = [
    (0,   99,  7410),
    (100, 149, 10550),
    (150, 199, 13830),
    (200, 249, 17510),
]

# Vika Fysikalske – ukjent for andre trinn, men 31-40 = 1 640 er kjent
# Fast: 199 abonnement + 790 rapporter
VIKA_TAB = [
    (0,  20,    0),   # antar ingen volumtillegg
    (21, 30,  860),
    (31, 40, 1640),
    (41, 50, 2290),   # estimert – kontroller med kontrakt
]

def lookup(table, calls):
    """Slå opp beløp i en trinntabell. Returnerer None hvis utenfor tabell."""
    for lo, hi, price in table:
        if lo <= calls <= hi:
            return price
    return None

def std(calls):
    return lookup(STDTAB, calls)

def ceg(calls):
    return lookup(CEGTAB, calls)

# ── Kundekonfigurasjon ────────────────────────────────────────────────────────
#
# Felt:
#   display   – visningsnavn i output
#   queues    – liste over Zisson-kønavn (nok at de STARTER MED strengen)
#   model     – prismodellkode (se beregn_belop)
#   ...modellspesifikke parametre...
#   manuell   – True hvis beløpet/deler av det må settes manuelt
#   notat     – merknad som vises i output

KUNDER = {

    # ── STANDARDTABELL ─────────────────────────────────────────────────────
    "to_clean": {
        "display": "2clean AS (Nordn)",
        "queues": ["Nordn (tidl. 2Clean)"],
        "model": "standard",
    },
    "advokat_olga": {
        "display": "Advokat Olga Halvorsen AS",
        "queues": ["Advokat Olga Halvorsen"],
        "model": "standard",
    },
    "advokat_sandra": {
        "display": "Advokat Sandra Latotinaite",
        "queues": ["Advokat Sandra Latotinaite"],
        "model": "standard",
    },
    "akkodis": {
        "display": "Akkodis Edge Norway AS",
        "queues": ["Akkodis"],
        "model": "standard",
    },
    "calpro": {
        "display": "Calpro AS",
        "queues": ["Calpro"],
        "model": "standard",
    },
    "coromatic": {
        "display": "Coromatic AS",
        "queues": ["Coromatic - Metric"],
        "model": "coromatic",
        "notat": "Volumtabell (Coromatic.xlsx). >599 anrop: 11228 + (anrop-599)×18. Statistikk til Linda Hertaas.",
    },
    "foma": {
        "display": "Foma Norge AS",
        "queues": ["Foma Service"],
        "model": "standard_tillegg",
        "tillegg": 199,
        "notat": "Standardtabell + kr 199 grunnpris tillegg",
    },
    "kardex": {
        "display": "Kardex Norge AS",
        "queues": ["Kardex"],
        "model": "standard",
    },
    "ktv_group": {
        "display": "KTV Group AS",
        "queues": ["KTV Group"],          # inkl. Kveld og helg
        "model": "standard_rabatt",
        "rabatt_pst": 10,
        "notat": "10 % fast rabatt på standardtabell",
    },
    "noah": {
        "display": "NOAH AS",
        "queues": ["Noah"],
        "model": "standard",
    },
    "norenco": {
        "display": "Norenco Norge AS",
        "queues": ["Norenco AS"],
        "model": "standard_tillegg",
        "tillegg": 199,
    },
    "senter_stress": {
        "display": "Senter for Stress og Traumepsykologi AS",
        "queues": ["Senter for Stress og Traumepsykologi"],
        "model": "standard",
    },
    "spraying_systems": {
        "display": "Spraying Systems Sverige AB",
        "queues": ["Spraying Systems Norge AS"],
        "model": "standard",
    },
    "standard_norge": {
        "display": "Standard Norge AS",
        "queues": ["Standard Norge", "Standard Online"],
        "model": "standard",
        "notat": "Standardtabell (Standard Online faktureres sammen med Standard Norge)",
    },
    "td_synnex": {
        "display": "TD Synnex Norway AS",
        "queues": ["TD Synnex"],
        "model": "standard",
    },
    "teva": {
        "display": "Teva AS (Teva-Actavis)",
        "queues": ["Teva"],
        "model": "standard_tillegg",
        "tillegg": 199,
    },
    "ultralydklinikken": {
        "display": "Ultralydklinikken AS",
        "queues": ["Ultralydklinikken"],
        "model": "standard",
    },
    "bodyfly": {
        "display": "Bodyfly Frogner AS",
        "queues": ["Bodyfly"],
        "model": "fast_pris",
        "fast": 1600,
        "notat": "Fast kr 1 600 inkl. 99 anrop",
    },

    # ── FAST PRIS + STYKKPRIS ──────────────────────────────────────────────
    "henjum": {
        "display": "Advokat Henjum AS",
        "queues": ["Advokat Henjum"],
        "model": "fast_stykkpris",
        "fast": 920, "inkl": 20, "per_anrop": 46,
    },
    "compass_group": {
        "display": "Compass Group Norge AS",
        "queues": ["Veidekke (3082)"],
        "model": "compass",
        "fast": 17000, "inkl": 800, "per_anrop_over800": 26,
        "email_pris": 10,
        "manuell": True,
        "notat": "Legg inn antall e-poster manuelt (se e-postlogg fra Veidekke)",
    },
    "frano": {
        "display": "Frano AS",
        "queues": ["Frano"],
        "model": "fast_stykkpris_alle",   # per_anrop gjelder alle anrop (ingen inkl)
        "fast": 195, "per_anrop": 29,
    },
    "funksjonell": {
        "display": "Funksjonell Medisinsk Klinikk AS",
        "queues": ["Funksjonellmedisinsk klinikk"],
        "model": "fast_pris",
        "fast": 840,
        "notat": "Betaler alltid kr 840 grunnpris uansett antall anrop",
    },
    "fysiodirekte": {
        "display": "Fysiodirekte AS",
        "queues": [],                   # ingen Zisson-kø funnet – alltid grunnpris
        "model": "fast_pris",
        "fast": 805,
        "notat": "Ingen Zisson-kø – betaler alltid kr 805 grunnpris",
    },
    "gulvfag": {
        "display": "Gulvfag Norge AS",
        "queues": ["Gulvfag Norge", "Adler parkettsliperi", "Nye Gulv Norge"],
        "model": "gulvfag",
        "fast": 700, "inkl": 20, "per_anrop": 36,
        "tillegg_per_kø": 139,
        "notat": "Kø 1 = Gulvfag (hoved). Kø 2+3 = Adler + Nye Gulv Norge à kr 139",
    },
    "gunvaldsen": {
        "display": "Gunvaldsen og sønn AS",
        "queues": ["Gunvaldsen og Sønn"],
        "model": "fast_stykkpris",
        "fast": 1180, "inkl": 40, "per_anrop": 27,
    },
    "ikt_nor": {
        "display": "IKT Nor AS",
        "queues": ["IKT-Nor"],
        "model": "fast_stykkpris",
        "fast": 860, "inkl": 20, "per_anrop": 42,
    },
    "invicta": {
        "display": "Invicta AS",
        "queues": ["Invicta"],
        "model": "fast_stykkpris",
        "fast": 2900, "inkl": 200, "per_anrop": 18,
    },
    "jcdecaux": {
        "display": "JCDecaux Norge AS",
        "queues": ["JCDecaux"],
        "model": "fast_stykkpris",
        "fast": 2000, "inkl": 20, "per_anrop": 40,
        "notat": "Ny pris fra juni 2026 (tidl. 4 480)",
    },
    "kontroll_elektro": {
        "display": "Kontroll Elektro AS",
        "queues": ["Kontroll Elektro"],
        "model": "fast_stykkpris",
        "fast": 860, "inkl": 20, "per_anrop": 42,
    },
    "meavia": {
        "display": "Meavia AS / Bekkestua Psykologen",
        "queues": ["Bekkestua psykologen"],
        "model": "fast_stykkpris_alle",
        "fast": 199, "per_anrop": 42,
    },
    "pc_vennen": {
        "display": "PC Vennen AS",
        "queues": ["PC Vennen (3795)"],
        "model": "fast_pris",
        "fast": 3010,
        "notat": "Fast kr 3 010 inkl. 149 anrop",
    },
    "positiv_trafikkskole": {
        "display": "Positiv Trafikkskole AS",
        "queues": ["Positiv Trafikkskole AS"],
        "model": "fast_stykkpris_alle",
        "fast": 515, "per_anrop": 42,
    },
    "prevas": {
        "display": "Prevas AS",
        "queues": ["Prevas"],
        "model": "fast_stykkpris_alle",
        "fast": 199, "per_anrop": 46,
    },
    "ringerike_elektro": {
        "display": "Ringerike Elektro AS",
        "queues": ["Ringerike Elektro"],
        "model": "fast_stykkpris",
        "fast": 764, "inkl": 20, "per_anrop": 30,
    },
    "safari_planter": {
        "display": "Safari Planter AS",
        "queues": ["Safari Planter"],
        "model": "fast_stykkpris_alle",
        "fast": 199, "per_anrop": 46,
    },
    "tropisk_design": {
        "display": "Tropisk Design AS",
        "queues": ["Tropisk Design"],
        "model": "fast_stykkpris",
        "fast": 880, "inkl": 20, "per_anrop": 45,
    },
    "zones": {
        "display": "Zones AS",
        "queues": ["Zones"],
        "model": "fast_pris",
        "fast": 820,
        "notat": "Back-up – betaler alltid kr 820 grunnpris",
    },
    "veidekke_more": {
        "display": "Veidekke Industri AS – Møre og Romsdal",
        "queues": ["Veidekke Industri Møre og Romsdal"],
        "model": "standard",
    },
    "veidekke_sogn": {
        "display": "Veidekke Industri AS – Sunnfjord/Sogn",
        "queues": ["Veidekke industri Asfalt"],
        "model": "standard",
    },

    # ── CEG-TABELL ─────────────────────────────────────────────────────────
    "ceg_norway": {
        "display": "CEG Norway AS",
        "queues": ["C-E-G Norway"],
        "model": "ceg",
    },
    "dental_sor": {
        "display": "Dental Sør AS",
        "queues": ["Dental Sør"],
        "model": "ceg",
    },
    "klinikk_onh": {
        "display": "Klinikk Øre-nese-hals AS",
        "queues": ["Klinikk Øre Nese Hals"],
        "model": "ceg",
    },
    "per_tverfjell": {
        "display": "Per Tverfjell Bilomsetning AS",
        "queues": ["Tverfjell Bilomsetning"],
        "model": "ceg",
        "notat": "OBS: Tripletex underfakturerte kr 390 i juni – kontroller neste faktura",
    },
    "vera_tank": {
        "display": "Vera Tank AS",
        "queues": ["Vera Tank"],
        "model": "ceg",
    },

    # NEMUS – én Zisson-kø for 5 klinikker; manuell fordeling
    "nemus": {
        "display": "NEMUS – 5 klinikker (Vestfold)",
        "queues": ["Nemus (4768)"],
        "model": "nemus_manuell",
        "manuell": True,
        "notat": "Fordel anrop per klinikk manuelt; CEG-tabell per klinikk. "
                 "Horten: fast 115+199. Larvik: fast 115+199. "
                 "Nøtterøy/Sandefjord/Tønsberg: fast 1500+199.",
    },

    # ── EGNE TABELLER ──────────────────────────────────────────────────────
    "ahlberg": {
        "display": "Ahlberg data (PC Vennen AS)",
        "queues": ["Ahlberg data"],
        "model": "ahlberg",
        "fast_inkl99": 3490, "over99": 32,
    },
    "allskog": {
        "display": "Allskog SA",
        "queues": ["Allskog"],           # inkl. tilbakeanrop-kø
        "model": "fast_pris",
        "fast": 15000,
        "notat": "Fast kr 15 000/mnd fra mars 2026 (ubegrenset + statistikk)",
    },
    "caverion": {
        "display": "Caverion Norge AS",
        "queues": ["Caverion"],
        "model": "tabell",
        "tabell": CAVERION_TAB,
        "notat": "Indeksregulert +3.2%. Send statistikk til Linda Hertaas og Lisa Johansson",
    },
    "efas": {
        "display": "EFAS (Eurofusion AS)",
        "queues": ["Efas"],
        "model": "efas",
        "fast_abo": 199,
        "tabell": EFAS_TAB,
    },
    "enrx": {
        "display": "ENRX AS",
        "queues": ["ENRX"],
        "model": "enrx",
        "tabell": ENRX_TAB,
        "over299": ENRX_OVER,
        "notat": "Husk engelsk fakturatekst",
    },
    "promon": {
        "display": "Promon AS",
        "queues": ["Promon"],
        "model": "tabell",
        "tabell": PROMON_TAB,
    },
    "ramboll": {
        "display": "Rambøll Norge AS",
        "queues": ["Rambøll"],
        "model": "ramboll",
        "base250_299": 4640, "over299_rate": 42, "terskel": 299,
        "notat": "Statistikk (kr 1 495) er sagt opp. Base kr 4 640 + kr 42/anrop over 299. "
                 "Faktura på engelsk, ref. 1353109",
    },
    "safenordic": {
        "display": "Safenordic AS",
        "queues": ["Safenordic", "Touchcom"],
        "model": "fast_pris",
        "fast": 1125,
        "notat": "Fast kr 1 125 inkl. 30 anrop (rollover 3 mnd). EHF-faktura",
    },
    "vika_fysikalske": {
        "display": "Vika Fysikalske Institutt AS",
        "queues": ["Vika Fysikalske Institutt"],
        "model": "vika",
        "fast_abo": 199, "rapport": 790,
        "tabell": VIKA_TAB,
        "notat": "199 fast + 790 rapporter + volumtabell",
    },

    # ── PER ANSATT ─────────────────────────────────────────────────────────
    "kroederen_elektro": {
        "display": "Krøderen Elektro AS",
        "queues": ["Krødern Elektro"],
        "model": "per_ansatt",
        "fast": 566, "per_ansatt": 114, "ansatte": 21, "per_anrop": 30,
        "notat": "Oppdater 'ansatte' ved endring i staben",
    },
    "menova": {
        "display": "Menova AS",
        "queues": ["Menova"],
        "model": "per_ansatt",
        "fast": 566, "per_ansatt": 56, "ansatte": 25, "per_anrop": 28,
        "notat": "Oppdater 'ansatte' ved endring i staben",
    },
    "willis": {
        "display": "Willis Towers Watson AS",
        "queues": ["WTW"],
        "model": "per_ansatt",
        "fast": 640, "per_ansatt": 118, "ansatte": 12, "per_anrop": 30,
        "notat": "Oppdater 'ansatte' ved endring i staben",
    },
    "hr_companies": {
        "display": "HR Companies AS",
        "queues": ["Advokatenes HR"],
        "model": "fast_stykkpris_alle",
        "fast": 873, "per_anrop": 30,
    },
    "print_supplies": {
        "display": "Print Supplies AS",
        "queues": ["Print Supplies"],
        "model": "fast_stykkpris_alle",
        "fast": 660, "per_anrop": 30,
    },
    "dekkteam": {
        "display": "Dekkteam Ringdekk AS",
        "queues": ["Ringdekk"],
        "model": "fast_stykkpris_alle",
        "fast": 1131, "per_anrop": 30,
    },

    # ── SPESIELLE ──────────────────────────────────────────────────────────
    "dal_media": {
        "display": "Dal Media (tidl. Digital Media Group AS)",
        "queues": ["Dal Media", "TDC Caravan og Service", "Saltando", "Victoria Vottestad"],
        "model": "dal_media",
        "tillegg_kø": ["TDC Caravan og Service", "Saltando", "Victoria Vottestad"],
        "tillegg_pris": 199,
        "notat": "Standardtabell på hoved-kø (Dal Media). 3 tilleggskøer à kr 199",
    },
    "fit4": {
        "display": "Fit4 AS",
        "queues": ["Fit4", "Helselaben"],
        "model": "fit4",
        "fast": 3292, "inkl": 100, "over_rate": 28,
        "email_pris": 0,   # sett manuelt per måned
        "manuell": True,
        "notat": "3 292 fast inkl. 100 + kr 28/over 100 + e-poster. Helselaben faktureres under samme avtale.",
    },
    "philips": {
        "display": "Philips Norge AS",
        "queues": ["Philips Norge"],
        "model": "fast_stykkpris",
        "fast": 8110, "inkl": 499, "per_anrop": 18,
        "rapport_tillegg": 2500,
        "notat": "8 110 base + 2 500 samtalelogg + 18/over 499",
    },
    "trafikkskolen_driver": {
        "display": "Trafikkskolen Driver AS",
        "queues": ["Trafikkskolen Driver"],
        "model": "fast_stykkpris",
        "fast": 6129, "inkl": 149, "per_anrop": 23,
        "notat": "Inkl. e-post og utgående anrop i pakken",
    },

    # ── KVARTALSFAKTURERING ────────────────────────────────────────────────
    "element_logic_no": {
        "display": "Element Logic AS",
        "queues": ["Element logic (11819)"],
        "model": "kvartal",
        "fast_kvartal": 8880,
        "notat": "Kvartalsvis forskudd. Neste faktura: 1. juli 2026",
    },
    "element_logic_se": {
        "display": "Element Logic Sweden AB",
        "queues": ["Element Logic Sverige"],
        "model": "kvartal",
        "fast_kvartal": 8880,
        "notat": "Kvartalsvis forskudd. Neste faktura: 1. juli 2026",
    },
    "enova_trondheim": {
        "display": "Enova Trondheim (Compass Group)",
        "queues": ["Enova Trondheim"],
        "model": "kvartal",
        "fast_kvartal": 5140,
        "notat": "Kvartalsvis forskudd. Neste faktura: 1. juli 2026",
    },
    "nyd": {
        "display": "NYD – Norsk Yrkesdykkerskole",
        "queues": ["NYD"],
        "model": "kvartal",
        "fast_kvartal": 10230,
        "notat": "Kvartalsvis forskudd. Neste faktura: 20. august 2026",
    },
    "nms": {
        "display": "Norsk Medisinsk Syklotronsenter",
        "queues": ["Norsk medisinsk Syklotronsenter"],
        "model": "nms_kvartal",
        "notat": "Kvartalsvis etterskudd. Bruk kvartalstotalen (ikke månedlig)",
    },

    "first_mover": {
        "display": "First Mover Group Norge AS",
        "queues": ["Relokator"],        # Zisson-kø heter fortsatt "Relokator" (gammelt navn)
        "model": "first_mover",
        "fast": 1240, "inkl": 99, "tilknytning": 199,
        "notat": "Månedspris kr 1 240 (0-99 anrop) + kr 199 ekstra tilknytning = kr 1 439. Over 99: avklar sats.",
    },
}

# ── Beregningsfunksjoner ──────────────────────────────────────────────────────

def beregn_belop(kunde_id, cfg, besvart_per_koe, **kwargs):
    """
    Returnerer (belop, anrop_hoved, notat_beregning)
    belop = None betyr manuell input nødvendig
    """
    model = cfg["model"]
    # Samle besvart for alle køer knyttet til kunden
    anrop = sum(v for k, v in besvart_per_koe.items()
                if any(k.startswith(q) for q in cfg.get("queues", [])))

    if model == "standard":
        p = std(anrop)
        return (p, anrop, f"Standardtabell {anrop} anrop → {p}")

    elif model == "standard_tillegg":
        p = std(anrop)
        if p is None: return (None, anrop, f"Over standardtabell ({anrop} anrop)")
        tot = p + cfg["tillegg"]
        return (tot, anrop, f"Standardtabell {p} + tillegg {cfg['tillegg']} = {tot}")

    elif model == "standard_rabatt":
        p = std(anrop)
        if p is None: return (None, anrop, "Over standardtabell")
        rabatt = round(p * cfg["rabatt_pst"] / 100)
        tot = p - rabatt
        return (tot, anrop, f"Standardtabell {p} − {cfg['rabatt_pst']}% ({rabatt}) = {tot}")

    elif model == "fast_pris":
        return (cfg["fast"], anrop, f"Fast pris {cfg['fast']}")

    elif model == "fast_stykkpris":
        over = max(0, anrop - cfg["inkl"])
        tot = cfg["fast"] + over * cfg["per_anrop"]
        rappport_til = cfg.get("rapport_tillegg", 0)
        tot += rappport_til
        txt = f"{cfg['fast']} fast inkl {cfg['inkl']}"
        if over:
            txt += f" + {over}×{cfg['per_anrop']} ({over*cfg['per_anrop']})"
        if rappport_til:
            txt += f" + {rappport_til} rapport"
        return (tot, anrop, txt)

    elif model == "fast_stykkpris_alle":
        tot = cfg["fast"] + anrop * cfg["per_anrop"]
        return (tot, anrop, f"{cfg['fast']} + {anrop}×{cfg['per_anrop']} = {tot}")

    elif model == "ceg":
        p = ceg(anrop)
        return (p, anrop, f"CEG-tabell {anrop} anrop → {p}")

    elif model == "tabell":
        p = lookup(cfg["tabell"], anrop)
        return (p, anrop, f"Volumtabell {anrop} anrop → {p}")

    elif model == "coromatic":
        if anrop > 599:
            p = 11228 + (anrop - 599) * COROMATIC_OVER_RATE
            return (p, anrop, f"Coromatic: 11228 + ({anrop}-599)×{COROMATIC_OVER_RATE} = {p}")
        p = lookup(COROMATIC_TAB, anrop)
        return (p, anrop, f"Coromatic volumtabell {anrop} anrop → {p}")

    elif model == "ahlberg":
        if anrop <= 99:
            return (cfg["fast_inkl99"], anrop, f"Fast {cfg['fast_inkl99']} (0-99)")
        else:
            over = anrop - 99
            tot = cfg["fast_inkl99"] + over * cfg["over99"]
            return (tot, anrop, f"{cfg['fast_inkl99']} + {over}×{cfg['over99']} = {tot}")

    elif model == "efas":
        p = lookup(cfg["tabell"], anrop)
        if p is None: return (None, anrop, f"Utenfor EFAS-tabell ({anrop})")
        tot = p + cfg["fast_abo"]
        return (tot, anrop, f"EFAS-tabell {p} + {cfg['fast_abo']} abo = {tot}")

    elif model == "enrx":
        p = lookup(cfg["tabell"], anrop)
        if p is not None:
            return (p, anrop, f"ENRX-tabell {anrop} → {p}")
        else:
            over = anrop - 299
            tot = 4640 + over * cfg["over299"]
            return (tot, anrop, f"ENRX: 4640 + {over}×{cfg['over299']} = {tot}")

    elif model == "ramboll":
        terskel = cfg["terskel"]
        over = max(0, anrop - terskel)
        tot = cfg["base250_299"] + over * cfg["over299_rate"]
        return (tot, anrop, f"Rambøll: {cfg['base250_299']} + {over}×{cfg['over299_rate']} = {tot}")

    elif model == "per_ansatt":
        tot = cfg["fast"] + cfg["ansatte"] * cfg["per_ansatt"] + anrop * cfg["per_anrop"]
        txt = (f"{cfg['fast']} + {cfg['ansatte']}×{cfg['per_ansatt']} "
               f"+ {anrop}×{cfg['per_anrop']} = {tot}")
        return (tot, anrop, txt)

    elif model == "dal_media":
        hoved_anrop = sum(v for k, v in besvart_per_koe.items()
                          if k.startswith("Dal Media"))
        p = std(hoved_anrop)
        n_tillegg = sum(1 for k in besvart_per_koe
                        if any(k.startswith(q) for q in cfg["tillegg_kø"]))
        tillegg_sum = n_tillegg * cfg["tillegg_pris"]
        tot = (p or 0) + tillegg_sum
        return (tot, anrop, f"Dal Media hoved-kø {hoved_anrop}→{p} + {n_tillegg}×{cfg['tillegg_pris']} = {tot}")

    elif model == "gulvfag":
        hoved_anrop = sum(v for k, v in besvart_per_koe.items()
                          if k.startswith("Gulvfag Norge"))
        tillegg_koer = [k for k in besvart_per_koe
                        if k.startswith("Adler") or k.startswith("Nye Gulv")]
        n_tillegg = len(tillegg_koer)
        over = max(0, hoved_anrop - cfg["inkl"])
        tot = cfg["fast"] + over * cfg["per_anrop"] + n_tillegg * cfg["tillegg_per_kø"]
        return (tot, anrop, f"Gulvfag: {cfg['fast']} + {over}×{cfg['per_anrop']} + {n_tillegg}×{cfg['tillegg_per_kø']} = {tot}")

    elif model == "fit4":
        over = max(0, anrop - cfg["inkl"])
        tot = cfg["fast"] + over * cfg["over_rate"] + kwargs.get("fit4_emails", 0) * cfg["email_pris"]
        return (tot, anrop, f"Fit4: {cfg['fast']} + {over}×{cfg['over_rate']} (manuell e-post legges til) = {tot}")

    elif model == "compass":
        over800 = max(0, anrop - cfg["inkl"])
        emails = kwargs.get("compass_emails", 0)
        tot = cfg["fast"] + over800 * cfg["per_anrop_over800"] + emails * cfg["email_pris"]
        txt = (f"{cfg['fast']} + {over800}×{cfg['per_anrop_over800']} "
               f"+ {emails} e-poster×{cfg['email_pris']} = {tot}")
        return (tot, anrop, txt)

    elif model == "vika":
        p = lookup(cfg["tabell"], anrop)
        tot = cfg["fast_abo"] + cfg["rapport"] + (p or 0)
        return (tot, anrop, f"Vika: {cfg['fast_abo']} abo + {cfg['rapport']} rapport + {p} vol = {tot}")

    elif model == "kvartal":
        return (0, anrop, f"Kvartalsvis – se notat")

    elif model == "nms_kvartal":
        return (None, anrop, "NMS kvartalsvis etterskudd – bruk kvartalstotal")

    elif model == "nemus_manuell":
        return (None, anrop, f"Nemus totalt {anrop} besvart – fordel per klinikk manuelt")

    elif model == "first_mover":
        # Standardtabell for volum + fast tillegg 1 240 + tilknytning 199
        p = std(anrop)
        if p is None:
            return (None, anrop, f"First Mover: {anrop} anrop – utenfor standardtabell")
        fast  = cfg.get("fast", 0)
        tilkn = cfg.get("tilknytning", 0)
        tot = p + fast + tilkn
        return (tot, anrop, f"First Mover: std {p} + {fast} fast + {tilkn} tilknytning = {tot}")

    elif model == "ukjent":
        return (None, anrop, f"UKJENT PRISMODELL – {anrop} anrop i Zisson")

    return (None, anrop, "Ikke håndtert")


# ── Les Zisson Excel ──────────────────────────────────────────────────────────

def les_zisson(excel_sti):
    import openpyxl
    wb = openpyxl.load_workbook(excel_sti)
    ws = wb["Kødetaljer første kø"]
    besvart = {}
    for row in ws.iter_rows(values_only=True):
        if row[1] != "Total":
            continue
        koenavn = str(row[0]) if row[0] else ""
        if koenavn in ("", "All queues"):
            continue
        b = row[5]
        besvart[koenavn] = int(b) if isinstance(b, (int, float)) else 0
    return besvart


# ── Bygg output Excel ─────────────────────────────────────────────────────────

def bygg_output(resultater, maaned_label, out_sti):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Fakturagrunnlag {maaned_label}"

    C_HDR = "1F4E79"
    C_OK  = "E2EFDA"
    C_MAN = "FFEB9C"
    C_OBS = "FFC7CE"
    C_WHT = "FFFFFF"
    C_ALT = "EBF3FB"

    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(bold=False, color="000000", size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)
    def bdr():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)
    def rgt(): return Alignment(horizontal="right",  vertical="center", wrap_text=True)

    COL_W = [5, 38, 10, 14, 14, 40]
    headers = ["#", "Kunde", "Besvart", "Beregnet ekskl.", "Status", "Merknader / beregning"]
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Tittelrad
    ws.row_dimensions[1].height = 26
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = fill(C_HDR)
        ws.cell(row=1, column=c).border = bdr()
    t = ws.cell(row=1, column=1,
        value=f"We4you AS – Fakturagrunnlag Sentralbord  |  {maaned_label}  |  Beregnet fra Zisson Interact")
    t.font = Font(bold=True, color="FFFFFF", size=11)
    t.alignment = lft()
    ws.merge_cells("A1:F1")

    # Headerrad
    ws.row_dimensions[2].height = 20
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.fill = fill(C_HDR); c.font = fnt(bold=True, color="FFFFFF")
        c.alignment = ctr(); c.border = bdr()

    nr = 0
    row = 3
    total = 0
    manuell_count = 0

    for kunde_id, res in resultater.items():
        cfg = KUNDER[kunde_id]
        belop, anrop, bereg_notat, notat_cfg = res

        if cfg.get("model") == "ukjent":
            bg = C_OBS
        elif cfg.get("manuell") or belop is None:
            bg = C_MAN
            manuell_count += 1
        elif (row - 2) % 2 == 0:
            bg = C_OK
        else:
            bg = C_ALT

        ws.row_dimensions[row].height = 28
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = fill(bg)
            ws.cell(row=row, column=c).border = bdr()

        nr += 1
        ws.cell(row=row, column=1, value=nr).alignment = ctr()
        ws.cell(row=row, column=1).font = fnt(color="595959")

        ws.cell(row=row, column=2, value=cfg["display"]).font = fnt(bold=True)
        ws.cell(row=row, column=2).alignment = lft()

        ca = ws.cell(row=row, column=3, value=anrop if anrop else "–")
        ca.alignment = ctr()
        ca.font = fnt(bold=True)

        if belop is not None:
            cb = ws.cell(row=row, column=4, value=belop)
            cb.number_format = '#,##0'
            cb.font = Font(bold=True, color="1F4E79", size=10)
            cb.alignment = rgt()
            total += belop
        else:
            ws.cell(row=row, column=4, value="MANUELL").alignment = ctr()
            ws.cell(row=row, column=4).font = fnt(bold=True, color="833C00")

        if cfg.get("manuell") or belop is None:
            status = "⚠ Manuell"
        elif cfg.get("model") == "ukjent":
            status = "🔴 Ukjent"
        else:
            status = "✓ Auto"
        ws.cell(row=row, column=5, value=status).alignment = ctr()
        ws.cell(row=row, column=5).font = fnt(italic=True,
            color="833C00" if "⚠" in status or "🔴" in status else "375623")

        merk = bereg_notat
        if notat_cfg:
            merk = f"{bereg_notat}  |  {notat_cfg}"
        ws.cell(row=row, column=6, value=merk).font = fnt(italic=True, color="595959")
        ws.cell(row=row, column=6).alignment = lft()

        row += 1

    # Totallinje
    ws.row_dimensions[row].height = 22
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = fill(C_HDR)
        ws.cell(row=row, column=c).border = bdr()
    tl = ws.cell(row=row, column=1, value=f"TOTAL BEREGNET ({maaned_label})")
    tl.font = Font(bold=True, color="FFFFFF", size=10)
    tl.alignment = lft()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    tv = ws.cell(row=row, column=4, value=total)
    tv.number_format = '#,##0'
    tv.font = Font(bold=True, color="FFC000", size=12)
    tv.alignment = rgt()
    tm = ws.cell(row=row, column=5, value=f"{manuell_count} manuelle poster")
    tm.font = fnt(italic=True, color="FFEB9C")
    tm.alignment = ctr()
    row += 1

    note = ws.cell(row=row, column=1,
        value="Grønn = automatisk beregnet  |  Gul = krever manuell input  |  Rød = ukjent prismodell")
    note.font = Font(italic=True, color="595959", size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    wb.save(out_sti)
    return total


# ── Hoved ─────────────────────────────────────────────────────────────────────

def main():
    # Parse argumenter
    args = sys.argv[1:]
    compass_emails = 0
    fit4_emails = 0
    zisson_sti = None

    skip_next = False
    for i, a in enumerate(args):
        if skip_next:
            skip_next = False
            continue
        if a == "--emails-compass" and i+1 < len(args):
            compass_emails = int(args[i+1])
            skip_next = True
        elif a == "--emails-fit4" and i+1 < len(args):
            fit4_emails = int(args[i+1])
            skip_next = True
        elif not a.startswith("--"):
            zisson_sti = a

    if not zisson_sti:
        # Finn nyeste Zisson-fil automatisk
        ZISSON_MAPPE = os.environ.get(
            "FAKTURAPLAN_DIR",
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data"),
        )
        filer = [f for f in os.listdir(ZISSON_MAPPE)
                 if f.startswith("We4you- Måned") and f.endswith(".xlsx")]
        if not filer:
            print("FEIL: Ingen Zisson-fil funnet. Oppgi sti som argument.")
            sys.exit(1)
        filer.sort(reverse=True)
        zisson_sti = os.path.join(ZISSON_MAPPE, filer[0])
        print(f"Bruker Zisson-fil: {filer[0]}")

    # Trekk ut månedslabel fra filnavn
    m = re.search(r"(\d{4}-\d{2}-\d{2}) - (.+)\.xlsx", os.path.basename(zisson_sti))
    if m:
        maaned_label = m.group(2).capitalize() + " " + m.group(1)[:4]
    else:
        maaned_label = date.today().strftime("%B %Y")

    print(f"Periode: {maaned_label}")
    print(f"Compass e-poster: {compass_emails}")
    print()

    # Les Zisson
    besvart_per_koe = les_zisson(zisson_sti)
    print(f"Leste {len(besvart_per_koe)} køer fra Zisson\n")

    # Beregn per kunde
    resultater = {}
    for kunde_id, cfg in KUNDER.items():
        belop, anrop, bereg_notat = beregn_belop(
            kunde_id, cfg, besvart_per_koe,
            compass_emails=compass_emails,
            fit4_emails=fit4_emails,
        )
        resultater[kunde_id] = (belop, anrop, bereg_notat, cfg.get("notat", ""))

    # Output
    OUT_MAPPE = os.environ.get(
        "FAKTURAPLAN_OUTDIR",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "fakturagrunnlag"),
    )
    os.makedirs(OUT_MAPPE, exist_ok=True)
    out_sti = os.path.join(OUT_MAPPE, f"Fakturagrunnlag_{maaned_label.replace(' ','_')}.xlsx")
    total = bygg_output(resultater, maaned_label, out_sti)

    print(f"{'Kunde':<42} {'Anrop':>7} {'Beregnet':>12}")
    print("-" * 65)
    manuell_poster = []
    for kunde_id, (belop, anrop, bereg_notat, notat_cfg) in resultater.items():
        cfg = KUNDER[kunde_id]
        if belop is None:
            manuell_poster.append((cfg["display"], anrop, bereg_notat))
            linje = "  ⚠  MANUELL"
        else:
            linje = f"  {belop:>12,}"
        print(f"{cfg['display']:<42} {anrop:>7} {linje}")

    print("-" * 65)
    print(f"{'TOTAL (automatisk)':<50} {total:>12,} kr")
    if manuell_poster:
        print(f"\n⚠  {len(manuell_poster)} poster krever manuell input:")
        for navn, anrop, notat in manuell_poster:
            print(f"   - {navn}: {notat}")
    print(f"\nLagret: {out_sti}")

if __name__ == "__main__":
    main()
