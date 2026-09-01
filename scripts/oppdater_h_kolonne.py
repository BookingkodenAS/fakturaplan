#!/usr/bin/env python3
"""
oppdater_h_kolonne.py  –  We4you AS sentralbord
================================================
Leser Zisson Interact maanedsstatistikk (Excel) og oppdaterer
H-kolonnen (Anrop besvart) i Prisoversikt-Sentralbord-*.xlsx.

Kjoer uten --skriv for aa se hva som vil endres (toerkjoering).
Kjoer med --skriv for aa faktisk lagre endringene.

OBS: Lukk Prisoversikten i Excel FOER du kjoerer med --skriv!

Bruk:
    python3 oppdater_h_kolonne.py [<zisson-fil>] [--skriv]

Eksempel:
    python3 oppdater_h_kolonne.py
    python3 oppdater_h_kolonne.py --skriv
"""

import sys
from pathlib import Path
import openpyxl

# Prosjektmappe
PROSJEKT = Path(r"C:\Users\hei\Claude\Projects\Fakturaplan")


def normaliser(tekst: str) -> str:
    """Fjern norske spesialtegn for robust matching (ø->o, æ->ae, å->aa)."""
    return (tekst.lower()
            .replace("ø", "o").replace("æ", "ae").replace("å", "aa")
            .replace("Ø", "o").replace("Æ", "ae").replace("Å", "aa"))


# Prioritert ko -> linje-mapping.
# Mer spesifikke moenstre OEVERST – sjekkes i rekkefolge.
# Linje None = NEMUS (spesialhandtert). Linje -1 = hopp over.
KO_LINJE = [
    # Spesifikke treff med ko-ID eller fullt navn
    ("Veidekke (3082)",                        27),   # Compass Group hovednummer
    ("Veidekke Industri More og Romsdal",      24),   # kø 3792
    ("Veidekke industri Asfalt",               25),   # kø 4509
    ("Victoria Vottestad",                      8),   # Saltando-underkø -> Dal Media
    ("Saltando (4520)",                          8),   # Saltando -> Dal Media
    ("Element Logic Sverige",                  10),   # Svensk – FØR norsk
    ("Element logic (11819)",                   9),   # Norsk
    ("Allskog (4833)",                         58),   # Ikke Allskog tilbakeanrop
    ("Allskog tilbakeanrop",                   -1),   # Hopp over
    ("PC Vennen (3795)",                       39),   # Ikke Ahlberg data
    ("We4you",                                 -1),   # Intern
    ("All queues",                             -1),   # Totalsum
    ("Hoyre",                                  -1),   # Ikke fakturerbar

    # Advokater
    ("Advokat Olga Halvorsen",                  2),
    ("Advokat Sandra Latotinaite",              3),
    ("Advokat Henjum",                         26),
    ("Advokatenes HR",                         71),   # HR Companies AS

    # Kjoente spesifikke koer
    ("Norsk medisinsk Syklotronsenter",        66),
    ("Senter for Stress og Traumepsykologi",   18),
    ("Funksjonellmedisinsk",                   30),
    ("Positiv Trafikkskole",                   41),
    ("Gass- og Pusteservice",                  13),
    ("Kontroll Elektro",                       37),
    ("Ringerike Elektro",                      75),
    ("Krodern Elektro",                        72),   # Krødern Elektro
    ("Print Supplies",                         74),
    ("Trafikkskolen Driver",                   76),
    ("Tverfjell Bilomsetning",                 55),
    ("Bekkestua psykologen",                   38),   # Meavia AS
    ("Klinikk Ore Nese Hals",                 49),   # Klinikk Øre Nese Hals

    # Kombinasjonskundar
    ("Standard Norge",                         20),   # + Standard Online summeres
    ("Standard Online",                        20),
    ("KTV Group",                              15),   # inkl. KTV Kveld og helg
    ("Norenco AS",                             17),
    ("Mekan",                                  17),   # Alias – telles med Norenco
    ("Relokator",                              11),   # First Mover Group

    # Dal Media-gruppen (alle 4 koer summeres for H-kolonnen)
    ("Dal Media",                               8),
    ("TDC Caravan",                             8),

    # Generelle koer
    ("TD Synnex",                              21),
    ("Dental Sor",                             48),   # Dental Sør
    ("Vera Tank",                              56),
    ("Enova Trondheim",                        65),
    ("Ahlberg data",                           57),
    ("Gulvfag Norge",                          32),
    ("Nye Gulv Norge",                         32),
    ("Adler parkettsliperi",                   32),
    ("Gunvaldsen og Sonn",                     33),   # og Sønn
    ("Safenordic",                             64),
    ("Touchcom",                               64),   # Alias – telles med Safenordic
    ("Spraying Systems",                       19),
    ("Tropisk Design",                         44),
    ("Ringdekk",                               69),   # Dekkteam Ringdekk
    ("Caverion",                               59),
    ("Invicta",                                35),
    ("Ramboll",                                63),   # Rambøll
    ("Prevas",                                 42),
    ("Promon",                                 62),
    ("Philips Norge",                          40),
    ("Coromatic",                               7),
    ("Akkodis",                                 4),
    ("Bodyfly",                                 5),
    ("Calpro",                                  6),
    ("Kardex",                                 14),
    ("Zones",                                  46),
    ("ENRX",                                   61),
    ("Efas",                                   60),
    ("Noah",                                   16),
    ("Nordn",                                   1),   # 2Clean AS
    ("Frano",                                  29),
    ("Ultralydklinikken",                      23),
    ("IKT-Nor",                                34),
    ("JCDecaux",                               36),
    ("Safari Planter",                         43),
    ("Menova",                                 73),
    ("Vika Fysikalske",                        45),
    ("Teva",                                   22),
    ("NYD",                                    67),
    ("WTW",                                    77),
    ("C-E-G Norway",                           47),
    ("Fit4",                                   70),
    ("Helselaben",                             70),   # Alias – telles med Fit4
    ("Foma Service",                           12),
    ("Nemus",                                None),   # Spesialhandtert via NEMUS-fil
    ("Allskog",                                58),   # Generisk fallback
]

# Linjer der H-kolonnen IKKE er anropsbasert
SKIP_LINJER = {28}   # Veidekke e-posthandtering


def match_linje(ko_navn: str):
    """Returner linje-nummer, -1 (hopp over), None (NEMUS), eller 'UKJENT'."""
    ko_norm = normaliser(ko_navn)
    for monster, linje in KO_LINJE:
        if normaliser(monster) in ko_norm:
            return linje
    return "UKJENT"


def finn_zisson_fil(mappe: Path, spesifisert=None) -> Path:
    if spesifisert:
        p = mappe / spesifisert
        if not p.exists():
            raise FileNotFoundError(f"Fant ikke: {p}")
        return p
    kandidater = sorted(
        list(mappe.glob("We4you* M*ned*.xlsx")),
        key=lambda f: f.stat().st_mtime, reverse=True
    )
    if not kandidater:
        raise FileNotFoundError(
            "Fant ingen Zisson-fil (We4you* Maaned*.xlsx) i mappen.\n"
            "Last opp manedsfilen til prosjektmappen og prov igjen."
        )
    return kandidater[0]


def finn_prisoversikt(mappe: Path) -> Path:
    kandidater = list(mappe.glob("Prisoversikt-Sentralbord-*.xlsx"))
    if not kandidater:
        raise FileNotFoundError("Fant ikke Prisoversikt-Sentralbord-*.xlsx")
    return sorted(kandidater)[-1]


def les_zisson_totaler(zisson_fil: Path) -> dict:
    """Les kun 'Total'-raden per ko. Returnerer {konavn: besvart_anrop}."""
    wb = openpyxl.load_workbook(str(zisson_fil), data_only=True, read_only=True)
    ws = wb["Kødetaljer første kø"]
    totaler = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ko_navn   = row[0]
        intervall = row[1]
        besvart   = row[5]  # Kolonne F = Besvart
        if str(intervall) == "Total" and ko_navn and isinstance(besvart, (int, float)):
            totaler[str(ko_navn)] = int(besvart)
    wb.close()
    return totaler


def beregn_linje_totaler(zisson_totaler: dict):
    linje_totaler = {}
    ukjente = []
    for ko_navn, besvart in zisson_totaler.items():
        linje = match_linje(ko_navn)
        if linje == -1:
            continue
        if linje == "UKJENT":
            ukjente.append(f"{ko_navn}: {besvart}")
            continue
        if linje is None or linje in SKIP_LINJER:
            continue
        linje_totaler[linje] = linje_totaler.get(linje, 0) + besvart
    return linje_totaler, ukjente


def les_radkart(pris_fil: Path) -> dict:
    wb = openpyxl.load_workbook(str(pris_fil), data_only=True, read_only=True)
    ws = wb.active
    radkart = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        a_val = row[0]
        if isinstance(a_val, (int, float)):
            linje = int(a_val)
            kunde = str(row[1])[:45] if len(row) > 1 and row[1] else ""
            h_val = row[7] if len(row) > 7 else None
            radkart[linje] = {"row": row_idx, "kunde": kunde, "h_na": h_val}
    wb.close()
    return radkart


def oppdater_fil(pris_fil: Path, endringer: list):
    wb = openpyxl.load_workbook(str(pris_fil))
    ws = wb.active
    for e in endringer:
        ws.cell(row=e["row"], column=8).value = e["ny_h"]
    wb.save(str(pris_fil))
    wb.close()


def oppdater_nemus(mappe: Path, total: int):
    kandidater = sorted(mappe.glob("NEMUS_Anrop_Fordeling*.xlsx"))
    if not kandidater:
        print("  !  NEMUS: Ingen NEMUS_Anrop_Fordeling*.xlsx funnet.")
        return
    nemus_fil = kandidater[-1]
    try:
        wb = openpyxl.load_workbook(str(nemus_fil))
        ws = wb.active
        ws["B5"] = total
        wb.save(str(nemus_fil))
        wb.close()
        print(f"  OK NEMUS: Oppdaterte B5 i {nemus_fil.name} -> {total} anrop")
        print("     -> Fyll inn fordeling per klinikk manuelt i NEMUS-filen (linje 50-54)")
    except PermissionError:
        print(f"  !  NEMUS: Filen er aapen i Excel – lukk den og prov igjen.")
    except Exception as ex:
        print(f"  !  NEMUS: Feil: {ex}")


def main():
    dry_run = "--skriv" not in sys.argv
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    spesifisert = args[0] if args else None

    print("=" * 62)
    print("  We4you – Oppdatering av anrop (H-kolonne) i Prisoversikt")
    print("=" * 62)

    try:
        zisson_fil = finn_zisson_fil(PROSJEKT, spesifisert)
        pris_fil   = finn_prisoversikt(PROSJEKT)
    except FileNotFoundError as e:
        print(f"\n  FEIL: {e}")
        sys.exit(1)

    print(f"  Zisson:       {zisson_fil.name}")
    print(f"  Prisoversikt: {pris_fil.name}")
    print(f"  Modus:        {'TOR-KJORING – ingen endringer lagres' if dry_run else 'SKRIV – endringer lagres'}")
    print()

    zisson_totaler = les_zisson_totaler(zisson_fil)
    linje_totaler, ukjente = beregn_linje_totaler(zisson_totaler)
    radkart = les_radkart(pris_fil)

    endringer = []
    ingen_endring = []
    for linje, ny_h in sorted(linje_totaler.items()):
        if linje not in radkart:
            print(f"  !  Linje {linje} finnes ikke i Prisoversikten")
            continue
        rad = radkart[linje]
        na_h = rad["h_na"]
        if str(na_h) != str(ny_h):
            endringer.append({
                "row": rad["row"], "linje": linje,
                "kunde": rad["kunde"], "na_h": na_h, "ny_h": ny_h,
            })
        else:
            ingen_endring.append(linje)

    if endringer:
        print(f"  {'Linje':>5}  {'Kunde':<42}  {'Na':>6}  ->  {'Ny':>6}")
        print("  " + "-" * 66)
        for e in endringer:
            print(f"  {e['linje']:>5}  {e['kunde']:<42}  {str(e['na_h']):>6}  ->  {e['ny_h']:>6}")
    else:
        print("  OK Ingen endringer – H-kolonnen er allerede a jour.")

    if ingen_endring:
        print(f"\n  Uendret ({len(ingen_endring)} linjer): {', '.join(str(l) for l in ingen_endring)}")

    nemus_total = zisson_totaler.get("Nemus (4768)")
    if nemus_total is not None:
        print(f"\n  NEMUS (ko 4768): {nemus_total} anrop totalt.")
        print("  -> Fordeling per klinikk (linje 50-54) oppdateres manuelt i NEMUS-filen.")

    if ukjente:
        print(f"\n  ! Ukjente koer (ikke i mapping – sjekk om ny kunde):")
        for u in ukjente:
            print(f"       {u}")

    if not dry_run:
        if endringer:
            try:
                oppdater_fil(pris_fil, endringer)
                print(f"\n  OK Lagret {len(endringer)} endringer i {pris_fil.name}")
            except PermissionError:
                print("\n  FEIL: Filen er aapen i Excel. Lukk Prisoversikten og kjoer paa nytt.")
                sys.exit(1)
        if nemus_total is not None:
            oppdater_nemus(PROSJEKT, nemus_total)
        print("\n  -> Aapne Prisoversikten i Excel – I-kolonnen beregnes automatisk.")
    elif endringer:
        print(f"\n  -> Kjoer med --skriv for aa lagre:")
        print(f"     python3 oppdater_h_kolonne.py --skriv")


if __name__ == "__main__":
    main()
